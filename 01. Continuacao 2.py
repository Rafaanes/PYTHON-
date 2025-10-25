# importar openpyxl
#lib que permite realizar formatações no excel
from openpyxl import load_workbook
import os

#item - caminho do Excel
arquivo_excel = r'C:\Rafaélla Sena - Python\Python 2\Python 2 - Clarify\Aluno\Aula_01\bases\Resultado_base_exe.xlsx'

#item 2 - carregar o arquivo Excel existente
# wb = Work Book
wb = load_workbook(arquivo_excel)

#item 3 - selecionar uma aba
#ws = Work Sheet
ws = wb['Relatório de vendas']

#item 4 - aplica a formatação de número(decimal com 2 casas) nas colunas F e G a partir da linha 2
#tradução: para cada linha dentro da planilha, inicie da linha 2. ws.max_row captura a quantidade de linhas preenchidas, ao somar mais 1, seu objetivo é descobrir a próxima linha vazia para saber o momento de parar/stop o for.
for row in range(2, ws.max_row + 1):
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].number_format = '#,##0.00'
    
try:
#item 5 - salvar as alterações no mesmo arquivo(sobreescrever)
    wb.save(arquivo_excel)
    os.startfile(arquivo_excel)

except: 
    print('Atenção! Arquivo aberto, ou verifique o erro')
    exit()
























































