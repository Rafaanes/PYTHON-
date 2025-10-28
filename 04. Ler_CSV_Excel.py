# importar lib
import pandas as pd
import os 

#importar a openpyxl para manipular o arquivo excel e aplicar formatação
from openpyxl import load_workbook
from openpyxl.styles import numbers

#item 1 - define os caminhos dos arquivos
origem_csv = r'C:\Rafaélla Sena - Python\Python 2\Python 2 - Clarify\Aluno\Aula_01\bases\Vendas Regionais (delimitadores).csv'

#item 2 - caminho Excel
arquivo_excel = r'C:\Rafaélla Sena - Python\Python 2\Python 2 - Clarify\Aluno\Aula_01\bases\Vendas Regionais (delimitadores).xlsx'

#item 3 - lê o CSV com o sep=';' e codificação 'latin-1'
df = pd.read_csv(origem_csv, sep = ';', encoding = 'latin-1')

#item 4 - converter a colunva de data para datetime e remove a parte do horário. dd/MM/yyyy
df['Data da Venda'] = pd.to_datetime(
    df['Data da Venda'], dayfirst=True).dt.normalize()

#item 5 - remover o simbolo [R$] e converter para float
df['Vendas'] = df ['Vendas'].replace('R\$\S?', '', regex=True).str.replace('.', '',regex=False).str.replace(',','.', regex=False).astype(float)

#item 6 - coluna [Comissões]
df['Comissões'] = df ['Comissões'].replace('R\$\S?', '', regex=True).str.replace('.', '',regex=False).str.replace(',','.', regex=False).astype(float)

#item 7 - abrir o Excel com o openxl para formatar a coluna [Data da Venda]
df.to_excel(arquivo_excel,index=False)

wb= load_workbook(arquivo_excel)
ws=wb.active #seleciona a planilha

#item 8 - encontra a coluna [Data da Venda] para aplicar formatação assumindo que está na coluna A,B,C.... Verificar dinamicamente o cabeçalho
header = [cell.value for cell in ws [1]] #tradução - para cada celula dentro da planilha número 1

#+1 por que col_idx no openpyxl começa em 1 
col_idx = header.index('Data da Venda') +1
col_vendas_idx = header.index('Vendas') +1
col_comissoes_idx = header.index('Comissões') +1

#item 9 - aplicar o formaro de data dd/MM/yyyy para toda a coluna de dados
for row in range(2,ws.max_row +1):
    cell = ws.cell(row=row, column = col_idx)
    cell.number_format = 'dd/MM/yyyy'

#item 10 - Vendas
for row in range(2, ws.max_row + 1):
    cell=ws.cell(row=row, column=col_vendas_idx)
    cell.number_format= '#,##0.00'

#item 11 - Comissões
for row in range(2,ws.max_row + 1):
    cell = ws.cell(row=row, column=col_comissoes_idx)
    cell.number_format = '#,##0.00'

#salvar o arquivo em excel
wb.save(arquivo_excel)

#abrir o arquivo
os.startfile(arquivo_excel)






















































































